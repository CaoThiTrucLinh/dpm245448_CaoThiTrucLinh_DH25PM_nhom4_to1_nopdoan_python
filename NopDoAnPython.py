import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import mysql.connector
from tkcalendar import DateEntry
from datetime import date
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import warnings
warnings.filterwarnings("ignore")

# -------------------- Thông tin DB - chỉnh theo máy bạn --------------------
def connect_db():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="qlnongduoc"
    )

# -------------------- Hàm tiện ích --------------------
def center_window(win, w=980, h=650):
    ws = win.winfo_screenwidth()
    hs = win.winfo_screenheight()
    x = (ws // 2) - (w // 2)
    y = (hs // 2) - (h // 2)
    win.geometry(f'{w}x{h}+{x}+{y}')

def format_currency_number(n):
    """Định dạng số nguyên thành chuỗi kiểu 70.000"""
    try:
        return f"{int(n):,}".replace(",", ".")
    except Exception:
        return "0"

def parse_currency_input(s):
    """Chuyển chuỗi nhập có thể có dấu chấm '70.000' thành int 70000"""
    if s is None: return 0
    if isinstance(s, (int, float)): return int(s)
    txt = str(s).strip()
    if txt == "":
        return 0
    # loại bỏ mọi ký tự không phải số
    cleaned = "".join(ch for ch in txt if ch.isdigit())
    try:
        return int(cleaned) if cleaned != "" else 0
    except ValueError:
        return 0

# -------------------- Tải dữ liệu lên Treeview --------------------
def load_data(filter_loai=None, filter_donvi=None):
    for i in tree.get_children():
        tree.delete(i)
    conn = connect_db()
    try:
        cur = conn.cursor()
        sql = "SELECT ma_thuoc, ten_thuoc, loai_thuoc, don_vi, so_luong, gia, ngay_nhap FROM thuoc WHERE 1=1"
        params = []
        if filter_loai and filter_loai != "Tất cả":
            sql += " AND loai_thuoc = %s"
            params.append(filter_loai)
        if filter_donvi and filter_donvi != "Tất cả":
            sql += " AND don_vi = %s"
            params.append(filter_donvi)
        cur.execute(sql, params)
        rows = cur.fetchall()
        for row in rows:
            ma, ten, loai, donvi, soluong, gia, ngay = row
            gia_display = format_currency_number(gia if gia is not None else 0)
            soluong_display = str(soluong) if soluong is not None else ""
            tree.insert("", tk.END, values=(ma, ten, loai, donvi, soluong_display, gia_display, str(ngay)))
    except Exception as e:
        messagebox.showerror("Lỗi CSDL", f"Không thể tải dữ liệu.\n{e}")
    finally:
        conn.close()

def clear_input():
    entry_ma.config(state='normal')
    entry_ma.delete(0, tk.END)
    entry_ten.delete(0, tk.END)
    cbb_loai.set("")
    entry_donvi.delete(0, tk.END)
    entry_soluong.delete(0, tk.END)
    entry_gia.delete(0, tk.END)
    date_entry.set_date(date.today())
    entry_ma.config(state='normal')

# -------------------- Thêm sản phẩm --------------------
def ThemSanPham():
    ma = entry_ma.get().strip()
    ten = entry_ten.get().strip()
    loai = cbb_loai.get().strip()
    donvi = entry_donvi.get().strip()
    soluong = entry_soluong.get().strip()
    gia_raw = entry_gia.get().strip()
    ngaynhap = date_entry.get_date()

    if ma == "" or ten == "" or loai == "":
        messagebox.showwarning("Thiếu dữ liệu", "Vui lòng nhập Mã thuốc, Tên thuốc và Loại thuốc")
        return

    try:
        soluong_val = int(soluong) if soluong != "" else 0
    except ValueError:
        messagebox.showwarning("Dữ liệu sai", "Số lượng phải là số nguyên")
        return

    gia_val = parse_currency_input(gia_raw)

    conn = connect_db()
    try:
        cur = conn.cursor()
        sql = """INSERT INTO thuoc
                 (ma_thuoc, ten_thuoc, loai_thuoc, don_vi, so_luong, gia, ngay_nhap)
                 VALUES (%s, %s, %s, %s, %s, %s, %s)"""
        val = (ma, ten, loai, donvi, soluong_val, gia_val, ngaynhap)
        cur.execute(sql, val)
        conn.commit()
        messagebox.showinfo("Thành công", "Thêm sản phẩm thành công")
        # reset filter combobox values list (nếu cần)
        refresh_filter_values()
        load_data(cbb_loc_loai.get(), cbb_loc_donvi.get())
        clear_input()
    except mysql.connector.IntegrityError:
        messagebox.showerror("Lỗi", f"Mã thuốc '{ma}' đã tồn tại!")
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

# -------------------- Xóa sản phẩm --------------------
def XoaSanPham():
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Chưa chọn", "Hãy chọn sản phẩm trên bảng để xóa")
        return
    ma = tree.item(selected)["values"][0]
    confirm = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa sản phẩm có mã {ma}?")
    if confirm:
        conn = connect_db()
        try:
            cur = conn.cursor()
            cur.execute("DELETE FROM thuoc WHERE ma_thuoc = %s", (ma,))
            conn.commit()
            refresh_filter_values()
            load_data(cbb_loc_loai.get(), cbb_loc_donvi.get())
            clear_input()
            messagebox.showinfo("Thành công", "Đã xóa sản phẩm")
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()

# -------------------- Sửa (đưa dữ liệu lên input) --------------------
def SuaSanPham(event=None):
    selected = tree.selection()
    if not selected:
        if event is None:
            messagebox.showwarning("Chưa chọn", "Hãy chọn sản phẩm để sửa")
        return
    values = tree.item(selected)["values"]
    entry_ma.config(state='readonly')
    entry_ma.delete(0, tk.END); entry_ma.insert(0, values[0])
    entry_ten.delete(0, tk.END); entry_ten.insert(0, values[1])
    cbb_loai.set(values[2])
    entry_donvi.delete(0, tk.END); entry_donvi.insert(0, values[3])
    entry_soluong.delete(0, tk.END); entry_soluong.insert(0, values[4])
    # giá đang hiển thị với dấu chấm; đưa thẳng vào ô
    entry_gia.delete(0, tk.END); entry_gia.insert(0, values[5])
    # ngày lưu dưới dạng 'YYYY-MM-DD' (string)
    try:
        date_entry.set_date(values[6])
    except Exception:
        date_entry.set_date(date.today())

# -------------------- Lưu sau khi sửa --------------------
def LuuSanPham():
    # Bắt buộc phải chọn Sửa trước (ma đang ở trạng thái readonly)
    if entry_ma.cget('state') != 'readonly':
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn một sản phẩm (Sửa) trước khi Lưu.")
        return

    ma = entry_ma.get().strip()
    ten = entry_ten.get().strip()
    loai = cbb_loai.get().strip()
    donvi = entry_donvi.get().strip()
    soluong = entry_soluong.get().strip()
    gia_raw = entry_gia.get().strip()
    ngaynhap = date_entry.get_date()

    try:
        soluong_val = int(soluong) if soluong != "" else 0
    except ValueError:
        messagebox.showwarning("Dữ liệu sai", "Số lượng phải là số nguyên")
        return

    gia_val = parse_currency_input(gia_raw)

    conn = connect_db()
    try:
        cur = conn.cursor()
        sql = """UPDATE thuoc
                 SET ten_thuoc=%s, loai_thuoc=%s, don_vi=%s, so_luong=%s, gia=%s, ngay_nhap=%s
                 WHERE ma_thuoc=%s"""
        val = (ten, loai, donvi, soluong_val, gia_val, ngaynhap, ma)
        cur.execute(sql, val)
        conn.commit()
        messagebox.showinfo("Thành công", "Cập nhật thông tin sản phẩm thành công")
        refresh_filter_values()
        load_data(cbb_loc_loai.get(), cbb_loc_donvi.get())
        clear_input()
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

# -------------------- Tìm kiếm --------------------
def TimKiem():
    search_win = tk.Toplevel(root)
    search_win.title("Tìm kiếm sản phẩm")
    search_win.geometry("380x150")
    tk.Label(search_win, text="Nhập mã, tên hoặc loại thuốc:").pack(pady=8)
    entry_search = tk.Entry(search_win, width=45); entry_search.pack(pady=4)
    def ThucHienTim():
        keyword = entry_search.get().strip()
        if keyword == "":
            messagebox.showwarning("Thông báo", "Vui lòng nhập từ khóa!")
            return
        for i in tree.get_children(): tree.delete(i)
        conn = connect_db()
        try:
            cur = conn.cursor()
            sql = """SELECT ma_thuoc, ten_thuoc, loai_thuoc, don_vi, so_luong, gia, ngay_nhap
                     FROM thuoc
                     WHERE ma_thuoc LIKE %s OR ten_thuoc LIKE %s OR loai_thuoc LIKE %s"""
            val = (f"%{keyword}%", f"%{keyword}%", f"%{keyword}%")
            cur.execute(sql, val)
            rows = cur.fetchall()
            if not rows:
                messagebox.showinfo("Kết quả", "Không tìm thấy sản phẩm nào.")
                load_data(cbb_loc_loai.get(), cbb_loc_donvi.get())
            else:
                for row in rows:
                    ma, ten, loai, donvi, soluong, gia, ngay = row
                    gia_display = format_currency_number(gia if gia is not None else 0)
                    soluong_display = str(soluong) if soluong is not None else ""
                    tree.insert("", tk.END, values=(ma, ten, loai, donvi, soluong_display, gia_display, str(ngay)))
            search_win.destroy()
        except Exception as e:
            messagebox.showerror("Lỗi", str(e))
        finally:
            conn.close()
    tk.Button(search_win, text="Tìm kiếm", command=ThucHienTim, width=15, bg="#9b59b6", fg="white").pack(pady=10)

# -------------------- Xuất Excel --------------------
def XuatExcel():
    conn = connect_db()
    try:
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                 title="Lưu file Excel")
        if not file_path:
            return
        cur = conn.cursor()
        cur.execute("SELECT ma_thuoc, ten_thuoc, loai_thuoc, don_vi, so_luong, gia, ngay_nhap FROM thuoc")
        rows = cur.fetchall()

        wb = Workbook(); ws = wb.active; ws.title = "Danh Sách Thuốc"
        headers = ["Mã thuốc", "Tên thuốc", "Loại thuốc", "Đơn vị", "Số lượng", "Giá (VND)", "Ngày nhập"]
        ws.append(headers)

        header_font = Font(bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in rows:
            # row: (ma, ten, loai, donvi, soluong, gia, ngay)
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.border = thin_border

        # Đặt format cột Giá là số
        try:
            col_gia = headers.index("Giá (VND)") + 1
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_gia)
                # Nếu là chuỗi dạng số, convert về số
                if isinstance(cell.value, str):
                    cleaned = "".join(ch for ch in cell.value if ch.isdigit())
                    if cleaned != "":
                        cell.value = int(cleaned)
                # Đặt định dạng số (Excel sẽ hiển thị theo locale)
                cell.number_format = '#,##0'
        except Exception:
            pass

        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 5, 40)

        wb.save(file_path)
        messagebox.showinfo("Thành công", f"Đã xuất file Excel tại:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
    finally:
        conn.close()

# -------------------- Lọc dữ liệu theo combobox --------------------
def LocDuLieu():
    loai = cbb_loc_loai.get()
    donvi = cbb_loc_donvi.get()
    load_data(loai, donvi)

# -------------------- Cập nhật danh sách giá trị cho combobox lọc (từ DB) --------------------
def refresh_filter_values():
    # Lấy danh sách loại và đơn vị hiện có trong DB để đưa vào combobox lọc
    conn = connect_db()
    try:
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT loai_thuoc FROM thuoc")
        loai_rows = [r[0] for r in cur.fetchall() if r[0]]
        cur.execute("SELECT DISTINCT don_vi FROM thuoc")
        donvi_rows = [r[0] for r in cur.fetchall() if r[0]]
        loai_values = ["Tất cả"] + sorted(loai_rows)
        donvi_values = ["Tất cả"] + sorted(donvi_rows)
        cbb_loc_loai['values'] = loai_values
        cbb_loc_loai.set("Tất cả")
        cbb_loc_donvi['values'] = donvi_values
        cbb_loc_donvi.set("Tất cả")
    except Exception:
        # nếu lỗi thì vẫn giữ giá trị mặc định
        cbb_loc_loai['values'] = ["Tất cả", "Thuốc trừ sâu", "Thuốc diệt cỏ", "Phân bón lá", "Thuốc trừ nấm", "Chất kích thích", "Khác"]
        cbb_loc_loai.set("Tất cả")
        cbb_loc_donvi['values'] = ["Tất cả", "gói", "chai", "lọ", "kg", "lít", "bao", "khác"]
        cbb_loc_donvi.set("Tất cả")
    finally:
        conn.close()

# -------------------- Giao diện chính --------------------
root = tk.Tk()
root.title("Quản Lý Cửa Hàng Thuốc Nông Dược")
center_window(root, 980, 650)
root.resizable(True, True)

lbl_title = tk.Label(root, text="QUẢN LÝ THUỐC NÔNG DƯỢC", font=("Arial", 20, "bold"), fg="#2c3e50")
lbl_title.pack(pady=8)

# Frame nhập liệu
frame_info = tk.Frame(root); frame_info.pack(pady=5, padx=10, fill="x")

tk.Label(frame_info, text="Mã thuốc:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
entry_ma = tk.Entry(frame_info, width=18); entry_ma.grid(row=0, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Tên thuốc:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
entry_ten = tk.Entry(frame_info, width=30); entry_ten.grid(row=0, column=3, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Loại thuốc:").grid(row=1, column=0, padx=5, pady=5, sticky="w")
cbb_loai = ttk.Combobox(frame_info, values=["Thuốc trừ sâu", "Thuốc diệt cỏ", "Phân bón lá", "Thuốc trừ nấm", "Chất kích thích", "Khác"], width=28, state="readonly")
cbb_loai.grid(row=1, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Đơn vị:").grid(row=1, column=2, padx=5, pady=5, sticky="w")
entry_donvi = tk.Entry(frame_info, width=15); entry_donvi.grid(row=1, column=3, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Số lượng:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_soluong = tk.Entry(frame_info, width=18); entry_soluong.grid(row=2, column=1, padx=5, pady=5, sticky="w")

tk.Label(frame_info, text="Giá (VND):").grid(row=2, column=2, padx=5, pady=5, sticky="w")
entry_gia = tk.Entry(frame_info, width=18); entry_gia.grid(row=2, column=3, padx=5, pady=5, sticky="w")
# placeholder ví dụ
entry_gia.insert(0, "70.000")

tk.Label(frame_info, text="Ngày nhập:").grid(row=3, column=0, padx=5, pady=5, sticky="w")
date_entry = DateEntry(frame_info, width=18, background="darkblue", foreground="white", date_pattern="yyyy-mm-dd")
date_entry.set_date(date.today())
date_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")

# Frame lọc (giống giao diện yêu cầu)
frame_filter = tk.Frame(root); frame_filter.pack(pady=6, padx=10, fill="x")
tk.Label(frame_filter, text="Lọc loại:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
cbb_loc_loai = ttk.Combobox(frame_filter, values=["Tất cả"], width=28, state="readonly")
cbb_loc_loai.grid(row=0, column=1, padx=5, pady=5, sticky="w")
tk.Label(frame_filter, text="Lọc đơn vị:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
cbb_loc_donvi = ttk.Combobox(frame_filter, values=["Tất cả"], width=20, state="readonly")
cbb_loc_donvi.grid(row=0, column=3, padx=5, pady=5, sticky="w")
tk.Button(frame_filter, text="Áp dụng lọc", command=LocDuLieu, width=15, bg="#95a5a6").grid(row=0, column=4, padx=10)

# Buttons chính
frame_btn = tk.Frame(root); frame_btn.pack(pady=6)
btn_width = 12
tk.Button(frame_btn, text="➕ Thêm", width=btn_width, command=ThemSanPham, bg="#2ecc71", fg="white").grid(row=0, column=0, padx=6)
tk.Button(frame_btn, text="💾 Lưu", width=btn_width, command=LuuSanPham, bg="#3498db", fg="white").grid(row=0, column=1, padx=6)
tk.Button(frame_btn, text="✍️ Sửa", width=btn_width, command=SuaSanPham, bg="#f1c40f").grid(row=0, column=2, padx=6)
tk.Button(frame_btn, text="❌ Hủy", width=btn_width, command=clear_input).grid(row=0, column=3, padx=6)
tk.Button(frame_btn, text="🗑️ Xóa", width=btn_width, command=XoaSanPham, bg="#e74c3c", fg="white").grid(row=0, column=4, padx=6)
tk.Button(frame_btn, text="🔍 Tìm Kiếm", width=btn_width, command=TimKiem, bg="#9b59b6", fg="white").grid(row=0, column=5, padx=6)
tk.Button(frame_btn, text="📊 Xuất Excel", width=btn_width, command=XuatExcel, bg="#1abc9c", fg="white").grid(row=0, column=6, padx=6)
tk.Button(frame_btn, text="🚪 Thoát", width=btn_width, command=root.quit).grid(row=0, column=7, padx=6)

tk.Label(root, text="Danh sách Thuốc", font=("Arial", 12, "bold")).pack(pady=6, anchor="w", padx=20)

# Treeview
columns = ("ma_thuoc", "ten_thuoc", "loai_thuoc", "don_vi", "so_luong", "gia", "ngay_nhap")
tree = ttk.Treeview(root, columns=columns, show="headings", height=15)
tree.heading("ma_thuoc", text="Mã thuốc"); tree.column("ma_thuoc", width=100, anchor="center")
tree.heading("ten_thuoc", text="Tên thuốc"); tree.column("ten_thuoc", width=260)
tree.heading("loai_thuoc", text="Loại"); tree.column("loai_thuoc", width=140, anchor="center")
tree.heading("don_vi", text="ĐVT"); tree.column("don_vi", width=80, anchor="center")
tree.heading("so_luong", text="Số lượng"); tree.column("so_luong", width=90, anchor="center")
tree.heading("gia", text="Giá (VND)"); tree.column("gia", width=120, anchor="center")
tree.heading("ngay_nhap", text="Ngày nhập"); tree.column("ngay_nhap", width=110, anchor="center")
tree.pack(padx=20, pady=5, fill="both", expand=True)
tree.bind("<Double-1>", SuaSanPham)

# Khởi tạo combobox lọc và load dữ liệu
refresh_filter_values()
load_data()

# Start
root.mainloop()
